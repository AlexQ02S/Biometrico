from flask import Flask, render_template, request, redirect, session, jsonify, send_file, url_for, flash, send_from_directory
import mysql.connector
import os
import json
from datetime import datetime, time, timedelta
import base64
import face_recognition
import requests
import cv2
import numpy as np
import pandas as pd
import ipaddress
import sys
import webbrowser
from threading import Timer
from threading import Lock
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfgen import canvas
from dotenv import load_dotenv
from werkzeug.middleware.proxy_fix import ProxyFix

from werkzeug.security import generate_password_hash, check_password_hash

def abrir_navegador():
    # Esto abrirá el navegador automáticamente en la PC del recepcionista
    webbrowser.open_new("http://127.0.0.1:5000")

buffer = BytesIO()

# Cargar variables de entorno desde .env (si existe)
DOTENV_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
load_dotenv(DOTENV_PATH)

# ===============================
# BASE PATH (LECTURA)
# ===============================
if getattr(sys, 'frozen', False):
    BASE_PATH = sys._MEIPASS
else:
    BASE_PATH = os.path.dirname(os.path.abspath(__file__))

TEMPLATES_DIR = os.path.join(BASE_PATH, "templates")
STATIC_DIR = os.path.join(BASE_PATH, "static")

# ===============================
# SAVE PATH (ESCRITURA REAL SEGURA)
# ===============================

# ===============================
# SAVE PATH (ESCRITURA REAL SEGURA)
# ===============================
if getattr(sys, 'frozen', False):
    # Si es EXE, guardamos en C:/Users/Usuario/Biometrico
    SAVE_PATH = os.path.join(os.path.expanduser("~"), "Biometrico")
else:
    # Si es script .py, usamos la carpeta actual
    SAVE_PATH = os.path.dirname(os.path.abspath(__file__))

# Definir carpetas de forma absoluta
CARPETA_STATIC = os.path.join(SAVE_PATH, "static")
CARPETA_FOTOS = os.path.join(CARPETA_STATIC, "fotos")
CARPETA_TEMP = os.path.join(CARPETA_STATIC, "temp")

# 🔥 Crear estructura completa (Aseguramos que existan todas las carpetas)
os.makedirs(CARPETA_FOTOS, exist_ok=True)
os.makedirs(CARPETA_TEMP, exist_ok=True)

app = Flask(
    __name__,
    template_folder=TEMPLATES_DIR,
    # El static_folder es la ruta física en el disco
    static_folder=CARPETA_STATIC,
    # El static_url_path es cómo se verá en el navegador (ej: /static/...)
    static_url_path='/static'
)

TRUST_PROXY_HEADERS = os.getenv("TRUST_PROXY_HEADERS", "false").strip().lower() == "true"
if TRUST_PROXY_HEADERS:
    # Permite obtener IP/protocolo reales cuando hay Nginx/Cloudflare delante.
    app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)

APP_SECRET_KEY = os.getenv("APP_SECRET_KEY")
if not APP_SECRET_KEY:
    raise RuntimeError("Falta la variable de entorno APP_SECRET_KEY")

app.secret_key = APP_SECRET_KEY
KNOWN_ENCODINGS = []
KNOWN_METADATA = []

# ===============================
# 🛡️ CONFIGURACIÓN Y MIDDLEWARE DE IP
# ===============================

ALLOWED_IPS_CSV = os.getenv(
    "ALLOWED_IPS_CSV",
    "45.173.230.31,45.4.203.191,200.10.15.20",
)
ALLOW_LOCALHOST_IPS = os.getenv("ALLOW_LOCALHOST_IPS", "true").strip().lower() == "true"

ALLOWED_IPS = [ip.strip() for ip in ALLOWED_IPS_CSV.split(",") if ip.strip()]
if ALLOW_LOCALHOST_IPS:
    ALLOWED_IPS.extend(["127.0.0.1", "::1"])

ALLOWED_NETWORKS_CSV = os.getenv(
    "ALLOWED_NETWORKS_CSV",
    "192.168.40.0/24,192.168.255.0/24",
)
ALLOWED_NETWORKS = []
for net in [n.strip() for n in ALLOWED_NETWORKS_CSV.split(",") if n.strip()]:
    try:
        ALLOWED_NETWORKS.append(ipaddress.ip_network(net, strict=False))
    except ValueError:
        print(f"Red inválida en ALLOWED_NETWORKS_CSV: {net}")

# 🔑 TOKENS SEGUROS CARGADOS DESDE .ENV
# Estructura JSON esperada en REMOTE_ACCESS_TOKENS_JSON:
# {"TOKEN_SECRETO": {"usuario_id": 1, "nombre": "Nombre"}}
remote_tokens_json = os.getenv("REMOTE_ACCESS_TOKENS_JSON", "{}")
try:
    TOKENS_ACCESO_REMOTO = json.loads(remote_tokens_json)
    if not isinstance(TOKENS_ACCESO_REMOTO, dict):
        TOKENS_ACCESO_REMOTO = {}
except json.JSONDecodeError:
    TOKENS_ACCESO_REMOTO = {}

MASTER_KEY = os.getenv("MASTER_KEY", "")

# Configuración de limpieza automática de fotos de asistencias
PHOTO_CLEANUP_ENABLED = os.getenv("PHOTO_CLEANUP_ENABLED", "true").strip().lower() == "true"
PHOTO_RETENTION_DAYS = int(os.getenv("PHOTO_RETENTION_DAYS", "14"))
PHOTO_CLEANUP_INTERVAL_DAYS = int(os.getenv("PHOTO_CLEANUP_INTERVAL_DAYS", "7"))
PHOTO_CLEANUP_STATE_FILE = os.path.join(SAVE_PATH, "photo_cleanup.last_run")
_PHOTO_CLEANUP_LOCK = Lock()


def obtener_ip_cliente():
    # En VPS con Nginx/Cloudflare, usamos cabeceras forward si están habilitadas.
    if TRUST_PROXY_HEADERS:
        cf_ip = request.headers.get("CF-Connecting-IP")
        if cf_ip and cf_ip.strip():
            return cf_ip.strip()

        xff = request.headers.get("X-Forwarded-For")
        if xff and xff.strip():
            # Tomamos la primera IP del encabezado (cliente real)
            return xff.split(",")[0].strip()

    return (request.remote_addr or "").strip()

@app.before_request
def restringir_por_ip():
    # 1. Permitir archivos estáticos (imágenes, CSS, etc.)
    if request.path.startswith('/static'):
        return

    # Ejecuta limpieza solo si corresponde por calendario
    ejecutar_limpieza_semanal_si_corresponde()

    # 2. VALIDAR SI VIENEN CON TOKEN REMOTO EN LA URL (ej: /?token=TOKEN_SEGURO)
    token_url = request.args.get("token")
    if token_url and token_url in TOKENS_ACCESO_REMOTO:
        # Autenticamos automáticamente en la sesión
        datos_user = TOKENS_ACCESO_REMOTO[token_url]
        session["remoto_autorizado"] = True
        session["usuario_remoto_id"] = datos_user["usuario_id"]
        return  # ¡Pasa directo!

    # 3. VALIDAR SI YA TIENEN LA SESIÓN DE TOKEN ACTIVA
    if session.get("remoto_autorizado"):
        return  # ¡Pasa directo!

    # 4. VALIDACIÓN TRADICIONAL DE IP (Para todos los demás)
    client_ip = obtener_ip_cliente()

    if client_ip in ALLOWED_IPS:
        return

    try:
        ip_obj = ipaddress.ip_address(client_ip)
        for network in ALLOWED_NETWORKS:
            if ip_obj in network:
                return
    except ValueError:
        pass

    # 5. DENEGAR ACCESO SI NO ES IP AUTORIZADA NI TIENE TOKEN
    return render_template("acceso_denegado.html", ip=client_ip), 403


# 4. RUTA ESPECIAL PARA VER LAS IMÁGENES
@app.route('/static/fotos/<path:filename>')
def custom_static(filename):
    return send_from_directory(CARPETA_FOTOS, filename)

# Ruta del modelo .dat
# Cambiado "datased" por "dataset" para que coincida con tu carpeta real
LANDMARKS_PATH = os.path.join(BASE_PATH, "dataset", "modelo", "shape_predictor_68_face_landmarks.dat")

# ===============================
# CONEXIÓN DB
# ===============================
def conectar():
   try:
        db_host = os.getenv("DB_HOST")
        db_user = os.getenv("DB_USER")
        db_password = os.getenv("DB_PASSWORD")
        db_name = os.getenv("DB_NAME")
        db_port = int(os.getenv("DB_PORT", "3306"))

        if not all([db_host, db_user, db_password, db_name]):
            raise RuntimeError(
                "Faltan variables de entorno de base de datos (DB_HOST, DB_USER, DB_PASSWORD, DB_NAME)"
            )

        conexion = mysql.connector.connect(
            host=db_host,
            user=db_user,
            password=db_password,
            database=db_name,
            port=db_port,
            connection_timeout=10,
            autocommit=True,
            use_pure=True
        )
        return conexion 
   except mysql.connector.Error as e:
        print("ERROR REAL DB:", e)
        return None

def cargar_rostros_en_memoria():
    global KNOWN_ENCODINGS, KNOWN_METADATA
    print("🔄 Cargando rostros en memoria...")
    
    con = conectar()
    if not con: return
    
    cursor = con.cursor(dictionary=True)
    cursor.execute("SELECT id, nombre, apellido, sede, foto FROM usuarios WHERE activo = 1")
    usuarios = cursor.fetchall()
    con.close()

    KNOWN_ENCODINGS = []
    KNOWN_METADATA = []
    

    for u in usuarios:
        ruta_foto = os.path.join(CARPETA_FOTOS, u["foto"])
        if os.path.exists(ruta_foto) and u["foto"] != "default.jpg":
            try:
                img = face_recognition.load_image_file(ruta_foto)
                encodings = face_recognition.face_encodings(img)
                if encodings:
                    KNOWN_ENCODINGS.append(encodings[0])
                    KNOWN_METADATA.append(u)
            except Exception as e:
                print(f"❌ Error cargando foto de {u['nombre']}: {e}")
    
    print(f"✅ {len(KNOWN_ENCODINGS)} rostros cargados correctamente.")
# ===============================
# PORTAL PRINCIPAL (NUEVO)
# ===============================
@app.route("/")
def portal():
    return render_template("portal.html")

# ===============================
# VERIFICAR PIN MAESTRO (NUEVO)
# ===============================
@app.route("/verificar_acceso_maestro", methods=["POST"])
def verificar_acceso_maestro():

    data = request.get_json()
    pin = data.get("pin")

    if pin == MASTER_KEY:

        session["terminal_activa"] = True

        return jsonify({
            "success": True,
            "redirect": "/index"
        })

    return jsonify({
        "success": False,
        "message": "Clave incorrecta"
    }), 401

# ===============================
# OBTENER HORA INTERNET
# ===============================
def obtener_hora_web():

    try:
        resp = requests.get(
            "http://worldtimeapi.org/api/timezone/America/Guayaquil",
            timeout=5
        )

        if resp.status_code == 200:

            data = resp.json()

            return datetime.fromisoformat(
                data['datetime'][:-6]
            )
    except Exception:
        pass

    return datetime.now()


def _leer_ultima_ejecucion_limpieza():
    if not os.path.exists(PHOTO_CLEANUP_STATE_FILE):
        return None

    try:
        with open(PHOTO_CLEANUP_STATE_FILE, "r", encoding="utf-8") as f:
            valor = f.read().strip()
        if not valor:
            return None
        return datetime.fromisoformat(valor)
    except Exception:
        return None


def _guardar_ultima_ejecucion_limpieza(fecha_hora):
    try:
        with open(PHOTO_CLEANUP_STATE_FILE, "w", encoding="utf-8") as f:
            f.write(fecha_hora.isoformat())
    except Exception as e:
        print("No se pudo guardar estado de limpieza:", e)


def limpiar_fotos_asistencia_antiguas():
    if not PHOTO_CLEANUP_ENABLED:
        return {"ok": True, "msg": "Limpieza deshabilitada", "eliminadas": 0, "actualizadas": 0}

    corte = (obtener_hora_web() - timedelta(days=PHOTO_RETENTION_DAYS)).date()
    con = conectar()
    if not con:
        return {"ok": False, "msg": "No se pudo conectar a BD para limpieza", "eliminadas": 0, "actualizadas": 0}

    eliminadas = 0
    actualizadas = 0

    try:
        cursor = con.cursor(dictionary=True)

        # Fotos de asistencias antiguas candidatas
        cursor.execute(
            """
            SELECT DISTINCT foto
            FROM asistencias
            WHERE fecha < %s
              AND foto IS NOT NULL
              AND foto <> ''
            """,
            (corte,),
        )
        fotos_asistencia = {
            row["foto"]
            for row in cursor.fetchall()
            if row.get("foto") and row["foto"] != "default.jpg"
        }

        if not fotos_asistencia:
            return {"ok": True, "msg": "No hay fotos antiguas para limpiar", "eliminadas": 0, "actualizadas": 0}

        # Evitar borrar fotos usadas como perfil de usuario
        cursor.execute(
            """
            SELECT DISTINCT foto
            FROM usuarios
            WHERE foto IS NOT NULL
              AND foto <> ''
            """
        )
        fotos_perfil = {
            row["foto"]
            for row in cursor.fetchall()
            if row.get("foto") and row["foto"] != "default.jpg"
        }

        fotos_a_limpiar = sorted(fotos_asistencia - fotos_perfil)
        if not fotos_a_limpiar:
            return {"ok": True, "msg": "No hay fotos limpiables (todas usadas por perfiles)", "eliminadas": 0, "actualizadas": 0}

        placeholders = ",".join(["%s"] * len(fotos_a_limpiar))

        # Primero desvinculamos en BD para no dejar referencias rotas
        try:
            cursor.execute(
                f"UPDATE asistencias SET foto = NULL WHERE fecha < %s AND foto IN ({placeholders})",
                [corte, *fotos_a_limpiar],
            )
        except mysql.connector.Error:
            # Fallback si la columna no permite NULL
            cursor.execute(
                f"UPDATE asistencias SET foto = '' WHERE fecha < %s AND foto IN ({placeholders})",
                [corte, *fotos_a_limpiar],
            )

        actualizadas = cursor.rowcount if cursor.rowcount and cursor.rowcount > 0 else 0
        con.commit()

        for nombre_foto in fotos_a_limpiar:
            ruta = os.path.join(CARPETA_FOTOS, nombre_foto)
            if os.path.exists(ruta):
                try:
                    os.remove(ruta)
                    eliminadas += 1
                except Exception as e:
                    print(f"No se pudo borrar {nombre_foto}: {e}")

        return {
            "ok": True,
            "msg": "Limpieza ejecutada",
            "eliminadas": eliminadas,
            "actualizadas": actualizadas,
        }

    except Exception as e:
        con.rollback()
        return {"ok": False, "msg": f"Error en limpieza: {str(e)}", "eliminadas": 0, "actualizadas": 0}
    finally:
        con.close()


def ejecutar_limpieza_semanal_si_corresponde(force=False):
    if not PHOTO_CLEANUP_ENABLED:
        return

    if not _PHOTO_CLEANUP_LOCK.acquire(blocking=False):
        return

    try:
        ahora = datetime.now()
        ultima = _leer_ultima_ejecucion_limpieza()
        due = (
            force
            or ultima is None
            or (ahora - ultima) >= timedelta(days=PHOTO_CLEANUP_INTERVAL_DAYS)
        )

        if not due:
            return

        resultado = limpiar_fotos_asistencia_antiguas()
        if resultado.get("ok"):
            _guardar_ultima_ejecucion_limpieza(ahora)
        print("Limpieza fotos:", resultado)
    finally:
        _PHOTO_CLEANUP_LOCK.release()

# ===============================
# GUARDAR IMAGEN BASE64
# ===============================
def guardar_imagen_base64(base64_string, ruta):
    try:
        if "," in base64_string:
            header, encoded = base64_string.split(",", 1)
        else:
            encoded = base64_string

        img_bytes = base64.b64decode(encoded)
        
        # En lugar de cv2.imwrite directamente, usamos este método 
        # que es compatible con rutas con tildes o espacios:
        nparr = np.frombuffer(img_bytes, np.uint8)
        img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        
        if img is None:
            raise ValueError("No se pudo decodificar la imagen")

        # Codificar y guardar usando numpy para evitar errores de ruta de Windows
        res, im_png = cv2.imencode(".jpg", img)
        with open(ruta, 'wb') as f:
            f.write(im_png.tobytes())

        print(f"Imagen guardada exitosamente en: {ruta}")

    except Exception as e:
        print(f"ERROR CRÍTICO AL GUARDAR: {e}")
        raise

# ===============================
# AUDITORIA
# ===============================
def registrar_auditoria(usuario_id, accion):

    try:

        con = conectar()
        cursor = con.cursor()

        ip = request.remote_addr

        cursor.execute(
            """
            INSERT INTO auditoria
            (usuario_id, accion, fecha, ip)
            VALUES (%s,%s,%s,%s)
            """,
            (usuario_id, accion, obtener_hora_web(), ip)
        )

        con.commit()
        con.close()

    except Exception as e:
        print("Error auditoria:", e)

# ===============================
# REGISTRAR ASISTENCIA BD
# ===============================
def registrar_asistencia_bd(usuario_id, tipo, nombre_foto):
    # Tipos válidos permitidos para evitar SQL Injection en la actualización
    TIPOS_PERMITIDOS = [
        "ingreso", "salida", 
        "inicio_almuerzo", "fin_almuerzo",
        "inicio_descanso", "fin_descanso", 
        "inicio_cena", "fin_cena"
    ]

    if tipo not in TIPOS_PERMITIDOS:
        return {
            "status": "error",
            "msg": f"Acción de registro inválida: {tipo}"
        }

    ahora = obtener_hora_web()
    con = conectar()
    if not con:
        return {"status": "error", "msg": "Error de conexión a la Base de Datos"}

    try:
        cursor = con.cursor(dictionary=True)

        cursor.execute(
            "SELECT * FROM asistencias WHERE usuario_id=%s AND fecha=%s",
            (usuario_id, ahora.date())
        )
        registro = cursor.fetchone()

        if registro:
            # Si ya registró esta acción hoy (ej: ya marcó 'inicio_descanso')
            if registro.get(tipo):
                return {
                    "status": "error",
                    "msg": f"{tipo.replace('_', ' ').capitalize()} ya fue registrado hoy"
                }

            # Actualizamos dinámicamente la columna que corresponda (Presencial o Híbrido)
            cursor.execute(
                f"UPDATE asistencias SET {tipo}=%s, foto=%s WHERE id=%s",
                (ahora.time(), nombre_foto, registro["id"])
            )

        else:
            # Primer registro del día para el usuario (Aplica para cualquier tipo de acción)
            ingreso = ahora.time() if tipo == "ingreso" else None
            salida = ahora.time() if tipo == "salida" else None
            
            # Modalidad Presencial
            inicio_almuerzo = ahora.time() if tipo == "inicio_almuerzo" else None
            fin_almuerzo = ahora.time() if tipo == "fin_almuerzo" else None
            
            # Modalidad Híbrido
            inicio_descanso = ahora.time() if tipo == "inicio_descanso" else None
            fin_descanso = ahora.time() if tipo == "fin_descanso" else None
            inicio_cena = ahora.time() if tipo == "inicio_cena" else None
            fin_cena = ahora.time() if tipo == "fin_cena" else None

            cursor.execute(
                """
                INSERT INTO asistencias
                (usuario_id, fecha, ingreso, salida,
                 inicio_almuerzo, fin_almuerzo,
                 inicio_descanso, fin_descanso,
                 inicio_cena, fin_cena, foto)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    usuario_id,
                    ahora.date(),
                    ingreso,
                    salida,
                    inicio_almuerzo,
                    fin_almuerzo,
                    inicio_descanso,
                    fin_descanso,
                    inicio_cena,
                    fin_cena,
                    nombre_foto
                )
            )

        con.commit()
        registrar_auditoria(usuario_id, f"Registro {tipo}")

        return {
            "status": "ok",
            "msg": f"{tipo.replace('_', ' ').capitalize()} registrado correctamente"
        }

    except Exception as e:
        con.rollback()
        return {
            "status": "error",
            "msg": f"Error al guardar en la base de datos: {str(e)}"
        }

    finally:
        # Cierre seguro garantizado de la conexión
        con.close()
# ===============================
# INDEX (BIOMÉTRICO)
# ===============================
@app.route("/index")
def index():

    if not session.get("terminal_activa"):
        return redirect("/")

    return render_template("index.html")

# ===============================
# LOGIN
# ===============================
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        usuario = request.form.get("usuario")
        password = request.form.get("password")

        con = conectar()
        if not con:
            return render_template(
                "login.html", error="Error de conexión con el servidor"
            )

        try:
            cursor = con.cursor(dictionary=True)
            cursor.execute(
                "SELECT * FROM usuarios WHERE usuario=%s AND rol='administrador'",
                (usuario,),
            )
            user = cursor.fetchone()
        finally:
            con.close()

        if (
            user
            and user.get("password")
            and check_password_hash(user["password"], password)
        ):
            session["user"] = user
            registrar_auditoria(user["id"], "Login")
            return redirect("/admin")

        return render_template("login.html", error="Credenciales incorrectas")

    return render_template("login.html")

# ===============================
# LOGOUT
# ===============================
@app.route("/logout")
def logout():

    if "user" in session:
        registrar_auditoria(session["user"]["id"], "Logout")

    session.clear()

    return redirect("/")
# ===============================
# PANEL ADMIN + FILTRO MES (CORREGIDO)
# ===============================

# ===============================
# PANEL ADMINISTRATIVO (COMPLETO)
# ===============================
@app.route("/admin")
def admin():
    if "user" not in session:
        return redirect("/")

    conexion = conectar()
    if not conexion:
        return "Error de conexión a la base de datos", 500

    # Fecha actual como valor por defecto
    hoy = datetime.now().strftime("%Y-%m-%d")

    # --------------------------------------------------
    # 1. PARÁMETROS GET (Soporta nombres simples y por modalidad)
    # --------------------------------------------------
    fecha_inicio = (
        request.args.get("fecha_inicio_general")
        or request.args.get("fecha_inicio_hibrido")
        or request.args.get("fecha_inicio")
        or hoy
    )

    fecha_fin = (
        request.args.get("fecha_fin_general")
        or request.args.get("fecha_fin_hibrido")
        or request.args.get("fecha_fin")
        or hoy
    )

    empleado_id = (
        request.args.get("empleado_general")
        or request.args.get("empleado_hibrido")
        or request.args.get("empleado")
    )

    try:
        cursor = conexion.cursor(dictionary=True)

        # --------------------------------------------------
        # 2. CONSULTA GENERAL DE ASISTENCIAS
        # --------------------------------------------------
        query = """
        SELECT 
            a.id,
            u.id as usuario_id,
            u.nombre,
            u.apellido,
            u.sede,
            u.area,
            u.foto as foto_usuario,
            u.foto as foto,
            a.fecha,
            a.ingreso,
            a.inicio_almuerzo,
            a.fin_almuerzo,
            a.inicio_descanso,
            a.fin_descanso,
            a.inicio_cena,
            a.fin_cena,
            a.salida,
            a.foto as foto_asistencia
        FROM asistencias a
        JOIN usuarios u ON u.id = a.usuario_id
        WHERE a.fecha BETWEEN %s AND %s AND u.activo = 1
        """

        valores = [fecha_inicio, fecha_fin]

        if empleado_id and empleado_id.strip() != "":
            query += " AND u.id = %s"
            valores.append(empleado_id)

        query += " ORDER BY a.fecha DESC, a.ingreso DESC"

        cursor.execute(query, valores)
        asistencias = cursor.fetchall()

        # --------------------------------------------------
        # 3. LISTADO DE EMPLEADOS ACTIVOS
        # --------------------------------------------------
        cursor.execute("""
            SELECT *
            FROM usuarios
            WHERE activo = 1
            ORDER BY nombre ASC, apellido ASC
        """)
        todos_los_empleados = cursor.fetchall()

        # --------------------------------------------------
        # 4. ESTADÍSTICAS DEL RANGO / DÍA
        # --------------------------------------------------
        cursor.execute(
            """
        SELECT 
            (SELECT COUNT(*) FROM usuarios WHERE activo = 1) as total_empleados,
            COUNT(a.id) as total_registros,
            SUM(
                CASE 
                    WHEN (a.inicio_almuerzo IS NOT NULL AND a.fin_almuerzo IS NULL) 
                    OR (a.inicio_descanso IS NOT NULL AND a.fin_descanso IS NULL) 
                    OR (a.inicio_cena IS NOT NULL AND a.fin_cena IS NULL) 
                    THEN 1 ELSE 0 
                END
            ) as en_receso,
            SUM(CASE WHEN a.inicio_descanso IS NOT NULL AND a.fin_descanso IS NULL THEN 1 ELSE 0 END) as en_descanso,
            SUM(CASE WHEN a.inicio_cena IS NOT NULL AND a.fin_cena IS NULL THEN 1 ELSE 0 END) as en_cena,
            SUM(CASE WHEN a.salida IS NOT NULL THEN 1 ELSE 0 END) as finalizados
        FROM asistencias a
        JOIN usuarios u ON u.id = a.usuario_id
        WHERE a.fecha BETWEEN %s AND %s AND u.activo = 1
        """,
            [fecha_inicio, fecha_fin],
        )

        stats = cursor.fetchone()

        if not stats or stats["total_empleados"] is None:
            stats = {
                "total_empleados": 0,
                "total_registros": 0,
                "en_receso": 0,
                "en_descanso": 0,
                "en_cena": 0,
                "finalizados": 0,
            }

    finally:
        conexion.close()

    # --------------------------------------------------
    # 5. RENDERIZADO DE PLANTILLA
    # --------------------------------------------------
    return render_template(
        "admin.html",
        asistencias=asistencias,
        empleados=todos_los_empleados,
        lista_empleados=todos_los_empleados,
        stats=stats,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        empleado_seleccionado=empleado_id,
        fecha_inicio_general=fecha_inicio,
        fecha_fin_general=fecha_fin,
        empleado_general=empleado_id,
        fecha_inicio_hibrido=fecha_inicio,
        fecha_fin_hibrido=fecha_fin,
        empleado_hibrido=empleado_id,
        user=session["user"],
    )
# ===============================
# EXPORTAR EXCEL (CORREGIDO Y BLINDADO)
# ===============================
@app.route("/exportar_excel")
def exportar_excel():
    con = conectar()
    if not con:
        return "Error de conexión a la base de datos", 500

    cursor = con.cursor(dictionary=True)

    empleado = request.args.get("empleado")
    fecha_inicio = request.args.get("fecha_inicio")
    fecha_fin = request.args.get("fecha_fin")
    sede = request.args.get("sede")

    if empleado in ["None", "", "undefined", None]:
        empleado = None
    if fecha_inicio in ["None", "", "undefined", None]:
        fecha_inicio = None
    if fecha_fin in ["None", "", "undefined", None]:
        fecha_fin = None
    if sede in ["None", "", "undefined", None]:
        sede = None

    query = """
    SELECT a.*, u.nombre, u.apellido, u.sede, u.area
    FROM asistencias a
    JOIN usuarios u ON u.id = a.usuario_id
    WHERE 1=1
    """

    valores = []

    if empleado:
        query += " AND u.id=%s"
        valores.append(empleado)

    if fecha_inicio and fecha_fin:
        query += " AND a.fecha BETWEEN %s AND %s"
        valores.append(fecha_inicio)
        valores.append(fecha_fin)

    if sede == "General":
        query += " AND u.sede != 'Hibrido'"
    elif sede == "Hibrido":
        query += " AND u.sede = 'Hibrido'"

    query += " ORDER BY a.fecha DESC, u.nombre"

    try:
        cursor.execute(query, tuple(valores))
        datos = cursor.fetchall()
    except Exception as e:
        print(f"Error en consulta Excel: {e}")
        return f"Error en la base de datos: {e}", 500
    finally:
        con.close()

    if not datos:
        return "No hay datos para exportar con los filtros seleccionados", 404

    # Generación de Excel con openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Asistencias"

    headers = [
        "ID",
        "Fecha",
        "Empleado",
        "Sede",
        "Área",
        "Ingreso",
        "Inicio Almuerzo",
        "Fin Almuerzo",
        "Inicio Descanso",
        "Fin Descanso",
        "Inicio Cena",
        "Fin Cena",
        "Salida",
    ]

    ws.append(headers)

    # Estilo de Encabezados
    header_fill = PatternFill(
        start_color="1F4E78", end_color="1F4E78", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Llenado de Filas
    for row in datos:
        nombre_completo = f"{row.get('nombre', '')} {row.get('apellido', '')}"
        ws.append(
            [
                row.get("id"),
                str(row.get("fecha", "")),
                nombre_completo,
                row.get("sede", ""),
                row.get("area", ""),
                str(row.get("ingreso", "") or "--:--"),
                str(row.get("inicio_almuerzo", "") or "--:--"),
                str(row.get("fin_almuerzo", "") or "--:--"),
                str(row.get("inicio_descanso", "") or "--:--"),
                str(row.get("fin_descanso", "") or "--:--"),
                str(row.get("inicio_cena", "") or "--:--"),
                str(row.get("fin_cena", "") or "--:--"),
                str(row.get("salida", "") or "--:--"),
            ]
        )

    # Ajuste automático de ancho de columnas
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"Reporte_Asistencias_{datetime.now().strftime('%Y%m%d')}.xlsx",
    )
# ===============================
# EXPORTAR PDF (COMPLETO Y BLINDADO)
# ===============================
@app.route("/exportar_pdf")
def exportar_pdf():
    con = conectar()
    if not con:
        return "Error de conexión a la base de datos", 500

    cursor = con.cursor(dictionary=True)

    tipo = request.args.get("tipo")
    empleado = request.args.get("empleado")
    fecha_inicio = request.args.get("fecha_inicio")
    fecha_fin = request.args.get("fecha_fin")

    if empleado in ["None", "", "undefined", None]:
        empleado = None
    if fecha_inicio in ["None", "", "undefined", None]:
        fecha_inicio = None
    if fecha_fin in ["None", "", "undefined", None]:
        fecha_fin = None

    query = """
    SELECT a.*, u.nombre, u.apellido, u.sede, u.area
    FROM asistencias a
    JOIN usuarios u ON u.id = a.usuario_id
    WHERE 1=1
    """

    valores = []

    if empleado:
        query += " AND u.id = %s"
        valores.append(empleado)

    if fecha_inicio:
        query += " AND a.fecha >= %s"
        valores.append(fecha_inicio)

    if fecha_fin:
        query += " AND a.fecha <= %s"
        valores.append(fecha_fin)

    query += " ORDER BY u.nombre, u.apellido, a.fecha DESC"

    try:
        cursor.execute(query, tuple(valores))
        datos = cursor.fetchall()
    except Exception as e:
        print("Error en consulta PDF:", e)
        return "Error en base de datos", 500
    finally:
        con.close()

    if not datos:
        return "No hay datos para generar el reporte PDF", 404

    def calc_neto(ing, sal, ini_a, fin_a):
        try:
            if not ing or not sal:
                return 0

            fmt = "%H:%M:%S"
            t_ing = datetime.strptime(str(ing), fmt)
            t_sal = datetime.strptime(str(sal), fmt)
            bruto = t_sal - t_ing
            pausa = timedelta(0)

            if ini_a and fin_a:
                pausa = datetime.strptime(str(fin_a), fmt) - datetime.strptime(
                    str(ini_a), fmt
                )

            return max(0, (bruto - pausa).total_seconds() / 3600)
        except Exception:
            return 0

    def calc_atraso_min(ing):
        if not ing:
            return 0
        try:
            h, m = map(int, str(ing).split(":")[:2])
            min_entrada = h * 60 + m
            return max(0, min_entrada - 485)  # 8:05 AM = 485 mins
        except Exception:
            return 0

    pdf_out = BytesIO()
    doc = SimpleDocTemplate(pdf_out, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    usuarios_reporte = {}
    for d in datos:
        key = (d["nombre"], d["apellido"], d["sede"], d["area"])
        if key not in usuarios_reporte:
            usuarios_reporte[key] = []
        usuarios_reporte[key].append(d)

    for (nom, ape, sede, area), registros in usuarios_reporte.items():
        elements.append(
            Paragraph(f"REPORTE INDIVIDUAL: {nom} {ape}", styles["Title"])
        )
        elements.append(
            Paragraph(
                f"<b>Sede:</b> {sede} | <b>Área:</b> {area}", styles["Normal"]
            )
        )
        elements.append(Spacer(1, 15))

        encabezados = [
            "Fecha",
            "Ingreso",
            "Salida",
            "Almuerzo",
            "Atraso",
            "Horas Netas",
        ]
        tabla_data = [encabezados]

        h_total = 0
        atrasos_count = 0

        for r in registros:
            min_atraso = calc_atraso_min(r["ingreso"])
            neto = calc_neto(
                r["ingreso"],
                r["salida"],
                r["inicio_almuerzo"],
                r["fin_almuerzo"],
            )

            if min_atraso > 0:
                atrasos_count += 1
            h_total += neto

            tabla_data.append(
                [
                    str(r["fecha"]),
                    str(r["ingreso"] or "--"),
                    str(r["salida"] or "--"),
                    (
                        f"{r['inicio_almuerzo']}-{r['fin_almuerzo']}"
                        if r["inicio_almuerzo"]
                        else "N/A"
                    ),
                    f"{min_atraso} min" if min_atraso > 0 else "Puntual",
                    f"{neto:.2f} hrs",
                ]
            )

        t = Table(tabla_data, colWidths=[70, 60, 60, 100, 70, 80])
        t.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4c1d95")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ]
            )
        )

        elements.append(t)
        elements.append(Spacer(1, 15))
        elements.append(
            Paragraph(
                f"<b>Resumen:</b> {atrasos_count} días con atraso | <b>Total:</b> {h_total:.2f} horas trabajadas.",
                styles["Normal"],
            )
        )
        elements.append(PageBreak())

    doc.build(elements)
    pdf_out.seek(0)

    return send_file(
        pdf_out,
        as_attachment=True,
        download_name=f"Reporte_Asistencias_{datetime.now().strftime('%Y%m%d')}.pdf",
        mimetype="application/pdf",
    )


# ===============================
# CREAR USUARIO
# ===============================
@app.route("/crear_usuario", methods=["POST"])
def crear_usuario():
    if "user" not in session:
        return redirect("/")

    nombre = request.form.get("nombre")
    apellido = request.form.get("apellido")
    sede = request.form.get("sede")
    area = request.form.get("area")
    rol = request.form.get("rol")
    usuario = request.form.get("usuario") if rol == "administrador" else None
    password = request.form.get("password") if rol == "administrador" else None

    if rol == "administrador" and password:
        password = generate_password_hash(password)

    foto_subida = request.files.get("foto_subida")
    foto_capturada = request.form.get("foto_capturada")

    os.makedirs(CARPETA_TEMP, exist_ok=True)
    os.makedirs(CARPETA_FOTOS, exist_ok=True)

    nombre_temp = f"verificar_{int(datetime.now().timestamp())}.jpg"
    temp_registro = os.path.join(CARPETA_TEMP, nombre_temp)

    try:
        if foto_subida and foto_subida.filename != "":
            foto_subida.save(temp_registro)
        elif foto_capturada:
            guardar_imagen_base64(foto_capturada, temp_registro)
        else:
            flash("Error: No se proporcionó ninguna imagen", "error")
            return redirect("/admin")
    except Exception as e:
        flash(f"Error al guardar la imagen: {str(e)}", "error")
        return redirect("/admin")

    if not os.path.exists(temp_registro):
        flash("Error interno: la imagen no se guardó correctamente", "error")
        return redirect("/admin")

    try:
        nueva_img = face_recognition.load_image_file(temp_registro)
        nueva_enc = face_recognition.face_encodings(nueva_img)
    except Exception as e:
        if os.path.exists(temp_registro):
            os.remove(temp_registro)
        flash(f"Error procesando la imagen: {str(e)}", "error")
        return redirect("/admin")

    if not nueva_enc:
        if os.path.exists(temp_registro):
            os.remove(temp_registro)
        flash("No se detectó un rostro claro en la imagen.", "error")
        return redirect("/admin")

    nueva_enc = nueva_enc[0]

    con = conectar()
    if not con:
        if os.path.exists(temp_registro):
            os.remove(temp_registro)
        flash("Error de conexión a la Base de Datos", "error")
        return redirect("/admin")

    try:
        cursor = con.cursor(dictionary=True)
        cursor.execute(
            "SELECT nombre, apellido, foto FROM usuarios WHERE activo = 1"
        )
        usuarios_existentes = cursor.fetchall()

        for u in usuarios_existentes:
            ruta_existente = os.path.join(CARPETA_FOTOS, u["foto"])
            if (
                not os.path.exists(ruta_existente)
                or u["foto"] == "default.jpg"
            ):
                continue

            try:
                img_existente = face_recognition.load_image_file(
                    ruta_existente
                )
                enc_existente = face_recognition.face_encodings(img_existente)

                if enc_existente:
                    match = face_recognition.compare_faces(
                        [enc_existente[0]], nueva_enc, tolerance=0.4
                    )
                    if match[0]:
                        if os.path.exists(temp_registro):
                            os.remove(temp_registro)
                        flash(
                            f"¡Esta persona ya está registrada como {u['nombre']} {u['apellido']}!",
                            "error",
                        )
                        return redirect("/admin")
            except Exception:
                continue

        nombre_foto = f"{nombre}_{int(datetime.now().timestamp())}.jpg"
        ruta_final = os.path.join(CARPETA_FOTOS, nombre_foto)

        os.replace(temp_registro, ruta_final)

        cursor.execute(
            """
            INSERT INTO usuarios 
            (nombre, apellido, sede, area, rol, usuario, password, foto)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                nombre,
                apellido,
                sede,
                area,
                rol,
                usuario,
                password,
                nombre_foto,
            ),
        )
        con.commit()
    finally:
        con.close()

    registrar_auditoria(session["user"]["id"], f"Crear usuario: {nombre}")
    cargar_rostros_en_memoria()

    flash(f"Usuario {nombre} registrado exitosamente", "success")
    return redirect("/admin")


# ===============================
# EDITAR USUARIO
# ===============================
@app.route("/editar_usuario", methods=["POST"])
def editar_usuario():
    if "user" not in session:
        return redirect("/")

    id_usuario = request.form.get("id")
    nombre = request.form.get("nombre")
    apellido = request.form.get("apellido")
    sede = request.form.get("sede")
    area = request.form.get("area")
    rol = request.form.get("rol")

    foto_subida = request.files.get("foto_subida_edit")
    foto_capturada = request.form.get("foto_capturada_edit")

    if rol not in ["empleado", "administrador"]:
        rol = "empleado"

    con = conectar()
    if not con:
        flash("Error de conexión a la base de datos", "error")
        return redirect("/admin")

    try:
        cursor = con.cursor(dictionary=True)
        cursor.execute("SELECT foto FROM usuarios WHERE id=%s", (id_usuario,))
        usuario_actual = cursor.fetchone()
        nombre_foto_db = (
            usuario_actual["foto"] if usuario_actual else "default.jpg"
        )

        if (foto_subida and foto_subida.filename != "") or foto_capturada:
            nuevo_nombre_foto = (
                f"{nombre}_{int(datetime.now().timestamp())}.jpg"
            )
            ruta_nueva = os.path.join(CARPETA_FOTOS, nuevo_nombre_foto)

            try:
                if foto_subida and foto_subida.filename != "":
                    foto_subida.save(ruta_nueva)
                elif foto_capturada:
                    guardar_imagen_base64(foto_capturada, ruta_nueva)

                if nombre_foto_db and nombre_foto_db != "default.jpg":
                    ruta_vieja = os.path.join(CARPETA_FOTOS, nombre_foto_db)
                    if os.path.exists(ruta_vieja):
                        os.remove(ruta_vieja)

                nombre_foto_db = nuevo_nombre_foto
            except Exception as e:
                print(f"Error al procesar foto en edición: {e}")

        if rol == "administrador":
            usuario = request.form.get("usuario")
            password = request.form.get("password")
            if password:
                password = generate_password_hash(password)
                cursor.execute(
                    """
                    UPDATE usuarios
                    SET nombre=%s, apellido=%s, sede=%s, area=%s,
                        rol=%s, usuario=%s, password=%s, foto=%s
                    WHERE id=%s
                """,
                    (
                        nombre,
                        apellido,
                        sede,
                        area,
                        rol,
                        usuario,
                        password,
                        nombre_foto_db,
                        id_usuario,
                    ),
                )
            else:
                cursor.execute(
                    """
                    UPDATE usuarios
                    SET nombre=%s, apellido=%s, sede=%s, area=%s,
                        rol=%s, usuario=%s, foto=%s
                    WHERE id=%s
                """,
                    (
                        nombre,
                        apellido,
                        sede,
                        area,
                        rol,
                        usuario,
                        nombre_foto_db,
                        id_usuario,
                    ),
                )
        else:
            cursor.execute(
                """
                UPDATE usuarios
                SET nombre=%s, apellido=%s, sede=%s, area=%s,
                    rol=%s, foto=%s
                WHERE id=%s
            """,
                (
                    nombre,
                    apellido,
                    sede,
                    area,
                    rol,
                    nombre_foto_db,
                    id_usuario,
                ),
            )

        con.commit()
    finally:
        con.close()

    cargar_rostros_en_memoria()
    registrar_auditoria(session["user"]["id"], "Editar usuario")

    return redirect("/admin")


# ===============================
# ELIMINAR USUARIO
# ===============================
@app.route("/eliminar_usuario/<int:id>", methods=["GET", "POST"])
def eliminar_usuario(id):
    if "user" not in session:
        return redirect("/")

    con = conectar()
    if not con:
        return redirect("/admin")

    try:
        cursor = con.cursor(dictionary=True)
        cursor.execute("SELECT foto FROM usuarios WHERE id=%s", (id,))
        user = cursor.fetchone()

        if user and user.get("foto") and user["foto"] != "default.jpg":
            ruta = os.path.join(CARPETA_FOTOS, user["foto"])
            if os.path.exists(ruta):
                try:
                    os.remove(ruta)
                except Exception:
                    pass

        cursor.execute("UPDATE usuarios SET activo = 0 WHERE id = %s", (id,))
        con.commit()
    finally:
        con.close()

    registrar_auditoria(
        session["user"]["id"], f"Deshabilitar usuario ID {id}"
    )
    cargar_rostros_en_memoria()

    return redirect("/admin")


# ===============================
# REGISTRAR ASISTENCIA (ENDPOINT)
# ===============================
@app.route("/registrar", methods=["POST"])
def registrar():
    data = request.json or {}

    usuario_id = data.get("usuario_id")
    tipo = data.get("tipo")
    foto = data.get("foto")

    if not usuario_id or not tipo or not foto:
        return jsonify(
            {"status": "error", "msg": "Datos incompletos para el registro"}
        )

    nombre_foto = f"{usuario_id}_{int(datetime.now().timestamp())}.jpg"

    guardar_imagen_base64(foto, os.path.join(CARPETA_FOTOS, nombre_foto))

    return jsonify(
        registrar_asistencia_bd(usuario_id, tipo, nombre_foto)
    )


# ===============================
# RECONOCER USUARIO
# ===============================
@app.route("/reconocer_usuario", methods=["POST"])
def reconocer_usuario():
    data = request.json or {}
    try:
        if "foto" not in data:
            return jsonify(
                {"status": "error", "msg": "No se envió la captura"}
            )

        img_bytes = base64.b64decode(data["foto"].split(",")[1])
        nparr = np.frombuffer(img_bytes, np.uint8)
        img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)

        small_img = cv2.resize(img, (0, 0), fx=0.25, fy=0.25)
        rgb_small_img = cv2.cvtColor(small_img, cv2.COLOR_BGR2RGB)

        face_locations = face_recognition.face_locations(
            rgb_small_img, model="hog"
        )
        unknown_encs = face_recognition.face_encodings(
            rgb_small_img, face_locations
        )

        if not unknown_encs:
            return jsonify(
                {"status": "error", "msg": "No se detectó rostro"}
            )

        face_to_check = unknown_encs[0]
        distancias = face_recognition.face_distance(
            KNOWN_ENCODINGS, face_to_check
        )

        if len(distancias) > 0:
            indice_mejor_match = np.argmin(distancias)
            min_distancia = distancias[indice_mejor_match]

            if min_distancia <= 0.45:
                u = KNOWN_METADATA[indice_mejor_match]
                precision = (1 - min_distancia) * 100
                return jsonify(
                    {
                        "status": "ok",
                        "usuario_id": u["id"],
                        "nombre": u["nombre"],
                        "apellido": u["apellido"],
                        "sede": u["sede"],
                        "precision": f"{precision:.2f}%",
                    }
                )

        return jsonify({"status": "error", "msg": "Rostro no reconocido"})
    except Exception as e:
        return jsonify({"status": "error", "msg": str(e)})


# ===============================
# INICIALIZACIÓN DE LA APLICACIÓN
# ===============================
if __name__ == "__main__":
    cargar_rostros_en_memoria()
    ejecutar_limpieza_semanal_si_corresponde(force=True)
    app_host = os.getenv("APP_HOST", "0.0.0.0")
    app_port = int(os.getenv("APP_PORT", "5000"))
    app_debug = os.getenv("APP_DEBUG", "false").strip().lower() == "true"
    app.run(host=app_host, port=app_port, debug=app_debug)